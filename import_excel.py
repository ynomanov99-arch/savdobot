import sqlite3
import os
import re
from openpyxl import load_workbook

DB_PATH = "costume_bot.db"
EXCEL_PATH = "ombor.xlsx"

def normalize_color(c):
    if not c:
        return ""
    c = str(c).strip().replace(" ", "")
    if "/" in c:
        parts = c.split("/")
        p1 = parts[0].lstrip("0")
        if not p1: p1 = "0"
        p2 = parts[1].lstrip("0")
        if not p2: p2 = "0"
        return f"{p1}/{p2}"
    return c.lstrip("0")

def normalize_model(m):
    if not m:
        return ""
    m = str(m).strip().lower()
    # Remove any extra spaces, plus signs, stars, or special characters
    m = re.sub(r'[^a-z0-9]', '', m)
    return m

def import_excel():
    if not os.path.exists(EXCEL_PATH):
        print(f"Error: '{EXCEL_PATH}' file not found!")
        print("Please place your Excel file in the bot folder named 'ombor.xlsx'.")
        return

    print("==================================================")
    print("EXCEL DATA INTEGRATION TO DATABASE")
    print("==================================================")
    print(f"File: {EXCEL_PATH}")
    print("Loading workbook...")

    try:
        wb = load_workbook(EXCEL_PATH, data_only=True)
        sheet = wb.active
    except Exception as e:
        print(f"Error loading Excel file: {e}")
        return

    print(f"Active Sheet: {sheet.title}, total rows: {sheet.max_row}")
    print("Loading database records...")

    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    
    # Fetch all database costumes to perform memory lookup matching
    cursor.execute("SELECT id, color_code, name, extra_note, sizes FROM costumes")
    db_costumes = cursor.fetchall()
    
    # Pre-process DB costumes with normalized keys
    db_lookup = {}
    for cid, c_code, c_name, c_extra, c_sizes in db_costumes:
        norm_code = normalize_color(c_code)
        norm_name = normalize_model(c_name)
        
        key = (norm_code, norm_name)
        if key not in db_lookup:
            db_lookup[key] = []
        db_lookup[key].append((cid, c_code, c_name, c_extra, c_sizes))

    print(f"Loaded {len(db_costumes)} database costumes.")
    print("Processing Excel rows and updating database...")

    success_count = 0
    total_processed = 0

    # Start from row 11 (which is after headers)
    for r in range(11, sheet.max_row + 1):
        model_val = sheet.cell(row=r, column=2).value
        drop_val = sheet.cell(row=r, column=3).value
        color_val = sheet.cell(row=r, column=6).value
        qty_val = sheet.cell(row=r, column=7).value
        location_val = sheet.cell(row=r, column=10).value

        # Skip empty rows
        if not color_val or not model_val:
            continue

        total_processed += 1

        xl_code = normalize_color(color_val)
        xl_model = normalize_model(model_val)
        
        qty = str(qty_val).strip() if qty_val is not None else "0"
        location = str(location_val).strip() if location_val is not None else "Ko'rsatilmagan"

        # Try exact lookup in memory first
        key = (xl_code, xl_model)
        matches = db_lookup.get(key)

        # Try partial matching if exact name match fails
        if not matches:
            for (db_code, db_model), db_rows in db_lookup.items():
                if db_code == xl_code and (xl_model in db_model or db_model in xl_model):
                    matches = db_rows
                    break

        if matches:
            # If multiple matches exist, pick the best one or update all matching variants
            for costume_id, _, _, _, _ in matches:
                cursor.execute(
                    "UPDATE costumes SET stock_qty = ?, warehouse_location = ? WHERE id = ?",
                    (qty, location, costume_id)
                )
            success_count += 1

    conn.commit()
    conn.close()

    print("==================================================")
    print("INTEGRATION COMPLETED SUCCESSFULLY!")
    print("==================================================")
    print(f"Total processed rows: {total_processed}")
    print(f"Successfully matched and updated rows: {success_count}")
    print("==================================================")

if __name__ == '__main__':
    import_excel()
